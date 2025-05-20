import pandas as pd
from datetime import date
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook
from openpyxl.styles import numbers

async def export_transactions_to_excel(pool, user_id: int) -> str:
    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT type, amount, category, date, description, 'Вручную' AS source
            FROM transactions
            WHERE user_id = $1

            UNION ALL

            SELECT 'expense' as type, ri.price as amount, ri.category, ri.purchase_date as date, '' as description, 'Чек' as source
            FROM receipt_items ri
            JOIN receipts r ON ri.receipt_id = r.id
            WHERE r.user_id = $1

            ORDER BY date DESC
        """, user_id)

    data = [dict(row) for row in rows]
    df = pd.DataFrame(data)
    df.columns = ["Тип", "Сумма", "Категория", "Дата", "Описание", "Источник"]

    temp_file = NamedTemporaryFile(delete=False, suffix=".xlsx")
    df.to_excel(temp_file.name, index=False)
    wb = load_workbook(temp_file.name)
    ws = wb.active

# Предполагаем, что "Сумма" — это второй столбец (B)
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):  # Столбец "Сумма"
        for cell in row:
            cell.number_format = '#,##0'  # Целое число без дробной части

    wb.save(temp_file.name)

    return temp_file.name
